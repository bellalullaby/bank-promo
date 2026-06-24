"""
从酱酱的 Excel 中提取银行 LOGO 并分析主色
输出：填好颜色的 Excel
"""
import sys
import io as _io
sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from PIL import Image
import io
import zipfile
import os
from collections import Counter

EXCEL_PATH = r'D:\2-营销活动\各银行LOGO文件\LOGO合集整理.xlsx'
OUTPUT_PATH = r'D:\2-营销活动\各银行LOGO文件\LOGO合集整理_填色.xlsx'

# 1. 从 Excel (ZIP) 中提取嵌入的图片
def extract_images_from_xlsx(xlsx_path):
    """Excel 文件本质是 ZIP，图片在 xl/media/ 里"""
    images = {}
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        for name in z.namelist():
            if 'media' in name.lower() and name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp')):
                images[name] = z.read(name)
    return images

# 2. 用 DISPIMG 公式中的 ID 匹配图片
def map_images_to_rows(wb, images):
    """通过 xl/drawings/ 中的关系匹配 DISPIMG ID → 图片文件"""
    mapping = {}

    # 读取 drawings 关系文件
    with zipfile.ZipFile(EXCEL_PATH, 'r') as z:
        # 列出所有文件
        drawing_files = [n for n in z.namelist() if 'drawing' in n.lower() and n.endswith('.xml')]

        for df in drawing_files:
            try:
                drawing_xml = z.read(df).decode('utf-8', errors='ignore')
            except:
                # try as binary
                continue

    # 另一种方法：直接按顺序匹配
    # DISPIMG ID 在 Excel 中是按图片嵌入顺序的
    # images 按文件名排序后顺序与行顺序一致
    sorted_images = sorted(images.keys())
    return sorted_images

# 3. 从图片提取主色
def get_dominant_color(img_data, k=1):
    """从图片数据中提取最主要的颜色"""
    img = Image.open(io.BytesIO(img_data))
    # 转 RGBA
    if img.mode != 'RGBA':
        img = img.convert('RGBA')

    # 缩小图片加速
    img = img.resize((100, 100))
    pixels = list(img.getdata())

    # 过滤掉透明像素和接近白/黑的像素
    filtered = []
    for r, g, b, a in pixels:
        if a < 128:  # 半透明以下跳过
            continue
        # 跳过接近纯白和纯黑的（背景/边框）
        if all(c > 240 for c in (r, g, b)):
            continue
        if all(c < 20 for c in (r, g, b)):
            continue
        # 跳过灰色（R≈G≈B）——通常是背景
        if max(r,g,b) - min(r,g,b) < 10:
            continue
        filtered.append((r, g, b))

    if not filtered:
        return (128, 128, 128)

    # 取最常见的颜色
    counter = Counter(filtered)
    dominant = counter.most_common(1)[0][0]
    return dominant

def rgb_to_hex(rgb):
    return '#{:02X}{:02X}{:02X}'.format(*rgb)

def get_light_version(rgb):
    """生成浅色版（用于卡片背景）"""
    # 混入大量白色
    r = min(255, rgb[0] + int((255 - rgb[0]) * 0.85))
    g = min(255, rgb[1] + int((255 - rgb[1]) * 0.85))
    b = min(255, rgb[2] + int((255 - rgb[2]) * 0.85))
    return (r, g, b)

def color_name(rgb):
    """给颜色起个中文名"""
    r, g, b = rgb
    if r > g and r > b:
        return '红色'
    elif g > r and g > b:
        return '绿色'
    elif b > r and b > g:
        return '蓝色'
    elif r > 150 and g > 100 and b < 80:
        return '橙色'
    elif r > 100 and g < 80 and b > 100:
        return '紫色'
    elif r > 120 and g > 100 and b < 60:
        return '棕金'
    else:
        return '深色'


print("=" * 50)
print("🏦 银行 LOGO 颜色提取器")
print("=" * 50)

# Step 1: 提取图片
print("\n📦 提取嵌入图片...")
images = extract_images_from_xlsx(EXCEL_PATH)
sorted_imgs = sorted(images.keys())
print(f"   找到 {len(sorted_imgs)} 张图片")

# Step 2: 读 Excel 行数据
print("\n📊 读取银行信息...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

banks = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    full_name = row[0].value  # A列
    short_name = row[1].value  # B列
    if full_name:
        banks.append({
            'full': full_name,
            'short': short_name,
            'row': row[0].row
        })

print(f"   找到 {len(banks)} 家银行")

# Step 3: 逐张分析颜色
print("\n🎨 分析 LOGO 主色...")

# 图片和银行按顺序对应
for i, (bank, img_name) in enumerate(zip(banks, sorted_imgs)):
    img_data = images[img_name]
    dominant = get_dominant_color(img_data)
    hex_color = rgb_to_hex(dominant)
    light_rgb = get_light_version(dominant)
    light_hex = rgb_to_hex(light_rgb)
    cname = color_name(dominant)

    print(f"   {i+1:2d}. {bank['short']:6s} → {hex_color} {cname}  (浅色: {light_hex})")

    # 写入 Excel
    row_num = bank['row']
    ws.cell(row=row_num, column=4, value=cname)   # D列：主色
    ws.cell(row=row_num, column=5, value=hex_color)  # E列：色号
    ws.cell(row=row_num, column=6, value=light_hex)  # F列：浅色版

# 加表头
ws.cell(row=1, column=6, value='浅色版')

# Step 4: 保存
print(f"\n💾 保存到: {OUTPUT_PATH}")
wb.save(OUTPUT_PATH)
print("✅ 完成！")
