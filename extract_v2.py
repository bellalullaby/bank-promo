"""
V2: 从各银行文件夹中找最干净的 LOGO 图片，用 K-means 聚色
"""
import sys, io as _io
sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from PIL import Image
import os, io
from collections import Counter

EXCEL_PATH = r'D:\2-营销活动\各银行LOGO文件\LOGO合集整理.xlsx'
LOGO_DIR = r'D:\2-营销活动\各银行LOGO文件'

# 银行文件夹名 → Excel 中对应简称的映射
# 从文件夹名和 Excel 数据建立对应关系

def find_logo_file(bank_folder):
    """在银行文件夹中找最干净的 LOGO 图片（优先 PNG，其次小尺寸 JPG）"""
    path = os.path.join(LOGO_DIR, bank_folder)
    if not os.path.isdir(path):
        return None

    files = []
    for f in os.listdir(path):
        full = os.path.join(path, f)
        if os.path.isfile(full) and f.lower().endswith(('.png', '.jpg', '.jpeg')):
            files.append((f, os.path.getsize(full), full))

    if not files:
        return None

    # 优先：文件名包含 'logo' 或只有 1 个文件
    logo_files = [(n,s,p) for n,s,p in files if 'logo' in n.lower() and s < 500000]

    if logo_files:
        # 取最小的（可能是最干净的 LOGO）
        logo_files.sort(key=lambda x: x[1])
        return logo_files[0][2]

    # 否则取最小的 PNG
    png_files = [(n,s,p) for n,s,p in files if n.lower().endswith('.png') and s < 500000]
    if png_files:
        png_files.sort(key=lambda x: x[1])
        return png_files[0][2]

    # 最后取最小的 JPG
    files.sort(key=lambda x: x[1])
    return files[0][2]

def extract_dominant_kmeans(img_data, n_colors=5):
    """用简化的颜色聚类找主色"""
    img = Image.open(io.BytesIO(img_data))
    if img.mode != 'RGBA':
        img = img.convert('RGBA')

    # 缩小
    img = img.resize((80, 80))
    pixels = list(img.getdata())

    # 过滤：去掉透明、白、黑、灰
    colored = []
    for r, g, b, a in pixels:
        if a < 128:
            continue
        if all(c > 240 for c in (r, g, b)):
            continue
        if all(c < 25 for c in (r, g, b)):
            continue
        # 灰色
        if max(r,g,b) - min(r,g,b) < 15:
            continue
        colored.append((r, g, b))

    if not colored:
        return (128, 128, 128)

    # 简化的"聚类"：按颜色分组
    # 把接近的颜色归到一起
    groups = []
    for c in colored:
        found = False
        for g in groups:
            # 如果颜色接近，归入该组
            if all(abs(c[i] - g['avg'][i]) < 40 for i in range(3)):
                g['sum'] = tuple(g['sum'][j] + c[j] for j in range(3))
                g['count'] += 1
                g['avg'] = tuple(int(g['sum'][j] / g['count']) for j in range(3))
                found = True
                break
        if not found:
            groups.append({'sum': c, 'count': 1, 'avg': c})

    # 按组大小排序
    groups.sort(key=lambda g: g['count'], reverse=True)

    if groups:
        return groups[0]['avg']
    return (128, 128, 128)

def process_folder_image(folder_name):
    """处理单个文件夹的 LOGO，返回主色"""
    img_path = find_logo_file(folder_name)
    if img_path:
        with open(img_path, 'rb') as f:
            img_data = f.read()
        return extract_dominant_kmeans(img_data)
    return None

def rgb_to_hex(rgb):
    return '#{:02X}{:02X}{:02X}'.format(*rgb)

def get_light_hex(rgb):
    r = min(255, rgb[0] + int((255 - rgb[0]) * 0.85))
    g = min(255, rgb[1] + int((255 - rgb[1]) * 0.85))
    b = min(255, rgb[2] + int((255 - rgb[2]) * 0.85))
    return '#{:02X}{:02X}{:02X}'.format(r, g, b)

def color_label(rgb):
    r, g, b = rgb
    if r > 160 and r > g*1.3 and r > b*1.3: return '红色'
    if g > 120 and g > r*1.2 and g > b*1.2: return '绿色'
    if b > 120 and b > r*1.2 and b > g*1.2: return '蓝色'
    if r > 140 and g > 80 and b < g*0.8: return '橙色'
    if r > 100 and b > 100 and g < min(r,b)*0.8: return '紫色'
    if r > 120 and g > 80 and b < 60: return '棕金'
    return '深色'


print("=" * 60)
print("V2: 从文件夹中找干净 LOGO + 颜色聚类")
print("=" * 60)

# 文件夹名到简称的映射（用于匹配 Excel 行）
folder_map = {
    '工行': '工商银行', '建行': '建设银行', '中行': '中国银行',
    '农行': '农业银行', '交行': '交通银行', '邮储': '邮储银行',
    '招商': '招商银行', '民生': '民生银行', '中信': '中信银行',
    '浦发': '浦发银行', '兴业': '兴业银行', '北京': '北京银行',
    '华夏': '华夏银行', '光大': '光大银行', '广发': '广发银行',
    '平安': '平安银行', '浙商': '浙商银行', '江西': '江西银行',
    '农商银行logo': '江西农商', '九江': '九江银行',
    '上饶': '上饶银行', '赣州': '赣州银行', '渤海': '渤海银行'
}

# 读 Excel
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

# 建立简称 → 行号 映射
short_to_row = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    short = row[1].value  # B列
    if short:
        short_to_row[short] = row[0].row

# 逐个文件夹处理
results = []
for folder_name, short_name in folder_map.items():
    print(f"\n🔍 {folder_name} → {short_name}...", end=' ')

    dominant = process_folder_image(folder_name)

    if dominant:
        hex_c = rgb_to_hex(dominant)
        light = get_light_hex(dominant)
        label = color_label(dominant)
        print(f"{hex_c} {label} | 浅色: {light}")

        # 写回 Excel
        if short_name in short_to_row:
            row_num = short_to_row[short_name]
            ws.cell(row=row_num, column=4, value=label)
            ws.cell(row=row_num, column=5, value=hex_c)
            ws.cell(row=row_num, column=6, value=light)
            results.append((short_name, hex_c, label))

        img_path = find_logo_file(folder_name)
        print(f"   📄 来源: {os.path.basename(img_path) if img_path else 'N/A'}")
    else:
        print("❌ 未找到可用图片")

ws.cell(row=1, column=6, value='浅色版(卡片用)')

OUTPUT = r'D:\2-营销活动\各银行LOGO文件\LOGO合集整理_填色V2.xlsx'
wb.save(OUTPUT)
print(f"\n{'=' * 60}")
print(f"✅ 保存到: {OUTPUT}")
print(f"   成功提取 {len(results)}/23 家银行颜色")
