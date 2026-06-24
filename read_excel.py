"""读取酱酱的银行LOGO Excel文件，查看结构"""
import openpyxl

wb = openpyxl.load_workbook(r'D:\2-营销活动\各银行LOGO文件\LOGO合集整理.xlsx')
ws = wb.active

print(f'Sheet: {ws.title}')
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
print()

# 打印所有行
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=False), 1):
    vals = [(c.value, c.column_letter) for c in row if c.value is not None]
    if vals:
        line = ' | '.join([f'{col}="{v}"' for v, col in vals])
    else:
        line = '[empty]'
    print(f'  Row {i}: {line}')
